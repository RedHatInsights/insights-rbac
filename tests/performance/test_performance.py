# Baseline tests for OCM performance
#
# 1 request for each tenant (to get tenant's groups)
# 1 request for each group (to get roles)
# 1 request for each tenant (to get the principles)
# 1 request for each principle in each org (to get the principles' groups)

# Populate the database with a large number of tenants, groups, principles, and roles
# ~20k requests = ~2k tenants ~10 principle per tenant
# optional req: 2 groups per principle, 5 roles per group with 10 permissions each

from http.client import REQUEST_TIMEOUT
from multiprocessing.pool import ThreadPool
import pdb
import time

import concurrent
from concurrent.futures import ThreadPoolExecutor, as_completed
from faker import Faker

from base64 import b64encode
from json import dumps as json_dumps
from urllib import request, response

from django.core.management.base import BaseCommand
import unittest

from django.test import SimpleTestCase, TestCase
from management.role.model import ExtRoleRelation, ExtTenant
from rbac.settings import DATABASES

from rest_framework import status

from api.models import Tenant, User
from api.serializers import create_tenant_name
from rbac.middleware import HttpResponseUnauthorizedRequest, IdentityHeaderMiddleware, TENANTS
from management.models import Access, Group, Permission, Principal, Policy, ResourceDefinition, Role

# for spreadsheeting
import openpyxl as xl

N_TENANTS = 2
GROUPS_PER_TENANT = 1

N = 1  # number of roles per group, number of principals per group
PRINCIPLES_PER_TENANT = 1

HEADERS = {
    "identity": {
        "account_number": "10001",
        "org_id": "11111",
        "type": "Associate",
                "user": {
                    "username": "user_dev",
                    "email": "user_dev@foo.com",
                    "is_org_admin": True,
                    "is_internal": True,
                    "user_id": "51736777",
                },
        "internal": {"cross_access": False},
    }
}

THREADS = 10

PATH = "ocm_performance.xlsx"

class OCMPerformanceTest(unittest.TestCase):
    def setUp(self):
        """Set up the test data."""
        print("Setting up test data...")

        tenants = []

        if ExtTenant.objects.filter(name="ocm").exists():
            ext_tenant = ExtTenant.objects.get(name="ocm")
        else:
            ext_tenant = ExtTenant.objects.create(
                name="ocm")

        def create_tenant(i):
            account = i
            t = Tenant.objects.create(
                org_id=i,
                account_id=account,
                tenant_name=f"ocm_acct{account}",
            )
            t.ready = True
            t.save()

            return t

        # create 2k tenants locally (so 2k different orgs)
        for i in range(N_TENANTS):
            tenants.append(create_tenant(i))

        def create_principal(tenant, i, j):
            Principal.objects.create(
                username=f"osetupcm_principal_{i}_{j}",
                tenant=tenant,
            )

        # for each org, create 10 principals
        for i in range(N_TENANTS):
            for j in range(PRINCIPLES_PER_TENANT):
                create_principal(tenants[i], i, j)

        def create_group(tenant, i, j):
            group = Group.objects.create(
                name=f"ocm_group_{i}_{j}",
                tenant=tenant,
            )

            policy = Policy.objects.create(
                name=f"ocm_policy_{i}_{j}",
                tenant=tenant,
            )

            roles = []
            # for each group, add "n" number of roles (should be OCM roles, so those with an external tenant)
            for k in range(N):
                role = Role.objects.create(
                    name=f"ocm_role_{i}_{j}_{k}",
                    tenant=tenant,
                )

                ExtRoleRelation.objects.create(
                    ext_id=f"ocm_role_relation_{i}_{j}_{k}",
                    ext_tenant=ext_tenant,
                    role=role,
                )

                policy.roles.add(role)

            policy.save()
            group.policies.add(policy)

            # for each group, assign "n" number of principals to the group
            for k in range(N):
                group.principals.add(Principal.objects.get(username=f"ocm_principal_{i}_{k}"))

            group.save()

        for i in range(N_TENANTS):
            for j in range(GROUPS_PER_TENANT):
                create_group(tenants[i], i, j)

        print("Finished setting up test data")

        self.wb = create_workbook()

    def test_seeding(self):
        """Test seeding."""
        self.assertEqual(Tenant.objects.count(), N_TENANTS)
        self.assertEqual(Principal.objects.count(), N_TENANTS * PRINCIPLES_PER_TENANT)
        self.assertEqual(Group.objects.count(), N_TENANTS * GROUPS_PER_TENANT)
        self.assertEqual(Policy.objects.count(), N_TENANTS * GROUPS_PER_TENANT)
        self.assertEqual(Role.objects.count(), N_TENANTS * GROUPS_PER_TENANT * N)

        print("Finished seeding test")

    def test_tenant_groups(self):
        """Test tenant groups with /integrations/tenant/{tenant_id}/groups/ endpoint."""

        tenants = Tenant.objects.exclude(tenant_name="public")
        # 1 request for each tenant (to get tenant's groups)

        name = "Tenant Groups"
        start = timerStart(name)

        def tenant_groups(self, tenant):
            response = self.client.get(
                f"/_private/api/v1/integrations/tenant/{tenant.org_id}/groups/?external_tenant=ocm",
                HEADERS,
                follow=True,
            )

            return response

        with ThreadPoolExecutor(max_workers=THREADS) as executor:
            futures = []
            for t in tenants:
                futures.append(executor.submit(tenant_groups, self=self, tenant=t))
            for future in as_completed(futures):
                response = future.result()
                self.assertEqual(response.status_code, status.HTTP_200_OK)
                self.assertEqual(response.data.get("meta").get("count"), GROUPS_PER_TENANT)

        # for t in tenants:
        #     response = self.client.get(
        #       f"/_private/api/v1/integrations/tenant/{t.org_id}/groups/?external_tenant=ocm",
        #         HEADERS,
        #         follow=True,
        #     )

        #     self.assertEqual(response.status_code, status.HTTP_200_OK)
        #     self.assertEqual(response.data.get("meta").get("count"), GROUPS_PER_TENANT)

        number_of_requests = N_TENANTS
        request_time, average = timerStop(start, number_of_requests)

        write_to_excel(
            self,
            name,
            "/api/v1/integrations/tenant/org_id/groups/",
            number_of_requests,
            request_time,
            average)

    def test_tenant_roles(self):
        """Test tenant roles with /integrations/tenant/{tenant_id}/roles/ endpoint."""
        tenants = Tenant.objects.exclude(tenant_name="public")
        # 1 request for each tenant (to get the tenant's roles)

        name = "Tenant Roles"
        start = timerStart(name)

        def tenant_roles(self, tenant):
            response = self.client.get(
                f"/_private/api/v1/integrations/tenant/{t.org_id}/roles/?external_tenant=ocm",
                HEADERS,
                follow=True,
            )

            return response

        with ThreadPoolExecutor(max_workers=1) as executor:
            futures = []
            for t in tenants:
                futures.append(executor.submit(tenant_roles, self=self, tenant=t))
            for future in as_completed(futures):
                response = future.result()
                self.assertEqual(response.status_code, status.HTTP_200_OK)
                self.assertEqual(response.data.get("meta").get("count"), N * GROUPS_PER_TENANT)

        number_of_requests = N_TENANTS
        request_time, average = timerStop(start, number_of_requests)

        write_to_excel(
            self,
            name,
            "/api/v1/integrations/tenant/{org_id}/roles/",
            number_of_requests,
            request_time,
            average
        )

    def test_group_roles(self):
        """Test group roles with /integrations/tenant/{tenant_id}/groups/{group_id}/roles/ endpoint."""
        # 1 request for each group (to get roles)

        tenants = Tenant.objects.exclude(tenant_name="public")

        pdb.set_trace()

        name = "Group Roles"
        start = timerStart(name)

        def group_roles(self, t, g):
            response = self.client.get(
                f"/_private/api/v1/integrations/tenant/{t.org_id}/groups/{g.uuid}/roles/?external_tenant=ocm",
                HEADERS,
                follow=True,
            )

            return response

        with ThreadPoolExecutor(max_workers=THREADS) as executor:
            futures = []
            for t in tenants:
                groups = Group.objects.filter(tenant=t)
                for g in groups:
                    futures.append(executor.submit(group_roles, self=self, t=t, g=g))
            for future in as_completed(futures):
                response = future.result()
                self.assertEqual(response.status_code, status.HTTP_200_OK)
                self.assertEqual(response.data.get("meta").get("count"), N)

        # for t in tenants:
        #     groups = Group.objects.filter(tenant=t)

        #     for g in groups:
        #         response = self.client.get(
        #             f"/_private/api/v1/integrations/tenant/{t.org_id}/groups/{g.uuid}/roles/?external_tenant=ocm",
        #             HEADERS,
        #             follow=True,
        #         )

        #         self.assertEqual(response.status_code, status.HTTP_200_OK)
        #         self.assertEqual(response.data.get("meta").get("count"), N)

        number_of_requests = N_TENANTS * GROUPS_PER_TENANT
        request_time, average = timerStop(start, number_of_requests)

        write_to_excel(
            self,
            name,
            "/api/v1/integrations/tenant/{org_id}/groups/{g_uuid}/roles/",
            number_of_requests,
            request_time,
            average
        )

    def test_principals_groups(self):
        """Test tenant principals groups with /integrations/tenant/{tenant_id}/principal/{principal_id}/groups/ endpoint."""
        # 1 request for each tenant (to get the principles)

        tenants = Tenant.objects.exclude(tenant_name="public")

        name = "Principals Groups"
        start = timerStart(name)

        def principals_groups(self, t, p):
            response = self.client.get(
                f"/_private/api/v1/integrations/tenant/{t.org_id}/principal/{p.username}/groups/?external_tenant=ocm",
                HEADERS,
                follow=True,
            )

            return response

        for t in tenants:
            principals = Principal.objects.filter(tenant=t)

            for p in principals:
                response = self.client.get(
                    f"/_private/api/v1/integrations/tenant/{t.org_id}/principal/{p.username}/groups/?external_tenant=ocm",
                    HEADERS,
                    follow=True,
                )

                self.assertEqual(response.status_code, status.HTTP_200_OK)
                self.assertEqual(response.data.get("meta").get("count"), GROUPS_PER_TENANT)

        number_of_requests = N_TENANTS * PRINCIPLES_PER_TENANT
        request_time, average = timerStop(start, N_TENANTS * PRINCIPLES_PER_TENANT)

        write_to_excel(
            self,
            name,
            "/api/v1/integrations/tenant/{org_id}/principal/{username}/groups/",
            number_of_requests,
            request_time,
            average
        )

    def test_principals_roles(self):
        """Test tenant principals roles with /integrations/tenant/{tenant_id}/principal/{principal_id}/groups/{group_id}/roles/ endpoint."""
        # 1 request for each tenant (to get the principles)
        tenants = Tenant.objects.exclude(tenant_name="public")

        name = "Principals Roles"
        start = timerStart(name)

        for t in tenants:
            principals = Principal.objects.filter(tenant=t)

            for p in principals:
                groups = Group.objects.filter(tenant=t)

                for g in groups:
                    response = self.client.get(
                        f"/_private/api/v1/integrations/tenant/{t.org_id}/principal/{p.username}/groups/{g.uuid}/roles/?external_tenant=ocm",
                        HEADERS,
                        follow=True,
                    )

                    self.assertEqual(response.status_code, status.HTTP_200_OK)
                    self.assertEqual(response.data.get("meta").get("count"), N)

        num_of_requests = N_TENANTS * PRINCIPLES_PER_TENANT * GROUPS_PER_TENANT

        request_time, average = timerStop(start, N_TENANTS * PRINCIPLES_PER_TENANT * GROUPS_PER_TENANT)

        write_to_excel(
            self,
            name,
            "/api/v1/integrations/tenant/{org_id}/principal/{username}/groups/{g_uuid}/roles/",
            num_of_requests,
            request_time,
            average
        )

    def test_full_sync(self):
        """Test full sync with /integrations/tenant/{tenant_id}/sync/ endpoint."""
        tenants = Tenant.objects.exclude(tenant_name="public")

        num_of_requests = N_TENANTS * (2 + GROUPS_PER_TENANT + PRINCIPLES_PER_TENANT)

        name = "Full Sync"
        start = timerStart(name)

        for t in tenants:
            # tenant groups, get orgs groups
            response = self.client.get(
                f"/_private/api/v1/integrations/tenant/{t.org_id}/groups/?external_tenant=ocm",
                HEADERS,
                follow=True,
            )
            self.assertEqual(response.status_code, status.HTTP_200_OK)
            self.assertEqual(response.data.get("meta").get("count"), GROUPS_PER_TENANT)

            groups = response.data.get("data")

            for g in groups:
                g_uuid = g.get("uuid")
                # tenant groups rolesk, get orgs groups roles
                response = self.client.get(
                    f"/_private/api/v1/integrations/tenant/{t.org_id}/groups/{g_uuid}/roles/?external_tenant=ocm",
                    HEADERS,
                    follow=True,
                )

                self.assertEqual(response.status_code, status.HTTP_200_OK)
                self.assertEqual(response.data.get("meta").get("count"), N)

            # no ocm specific endpoint to get principals per tenant

            # make header with tenant specified
            head = {
                "identity": {
                    "account_number": t.account_id,
                    "org_id": t.org_id,
                    "type": "Associate",
                            "user": {
                                "username": "user_dev",
                                "email": "user_dev@foo.com",
                                "is_org_admin": True,
                                "is_internal": True,
                                "user_id": "51736777",
                            },
                    "internal": {"cross_access": False},
                }
            }

            response = self.client.get(
                f"/api/rbac/v1/principals/",
                **head,
                follow=True,
            )

            self.assertEqual(response.status_code, status.HTTP_200_OK)
            # self.assertEqual(len(response.data.get("data")), PRINCIPLES_PER_TENANT)
            # self.assertEqual(response.data.get("meta").get("count"), PRINCIPLES_PER_TENANT)

            # principals = response.data.get("data")
            principals = Principal.objects.filter(tenant=t)
            for p in principals:
                # p_username = p.get("username")
                p_username = p.username
                response = self.client.get(
                    f"/_private/api/v1/integrations/tenant/{t.org_id}/principal/{p_username}/groups/?external_tenant=ocm",
                    HEADERS,
                    follow=True,
                )

                self.assertEqual(response.status_code, status.HTTP_200_OK)
                self.assertEqual(response.data.get("meta").get("count"), GROUPS_PER_TENANT)

        request_time, average = timerStop(start, num_of_requests)

        write_to_excel(self, name, "", num_of_requests, request_time, average)

    @classmethod
    def tearDownClass(self):
        """Clean up the test."""

        Group.objects.all().delete()
        Role.objects.all().delete()
        Policy.objects.all().delete()
        Principal.objects.all().delete()
        Tenant.objects.all().delete()
        ExtTenant.objects.all().delete()
        ExtRoleRelation.objects.all().delete()
        super().tearDownClass()

# ---------------------------
# Header building
# ---------------------------


def create_customer_data(account, tenant, org_id):
    """Create customer data."""
    customer = {"account_id": account, "tenant_name": tenant, "org_id": org_id}
    return customer


def create_user_data(fake):
    """Create user data."""
    user_data = {"username": fake.user_name(), "email": fake.email()}
    return user_data

# ---------------------------
# A couple of logging helpers
# ---------------------------


def timerStart(test_title):
    """Start timer."""
    start = time.perf_counter()

    print(f"Starting test for {test_title}...")

    return start


def timerStop(start, num_requests):
    """Stop timer."""
    request_time = time.perf_counter() - start

    average = request_time / num_requests

    print(f"Total request time: {request_time} seconds")
    print("Average time: {} seconds".format(average))

    return request_time, average

def create_workbook():
    # Create spreadsheet
    wb = xl.Workbook()
    ws = wb.active

    c1 = ws.cell(row=1, column=1)
    c1.value = "Test Name"

    c2 = ws.cell(row=1, column=2)
    c2.value = "Request URL"

    c3 = ws.cell(row=1, column=3)
    c3.value = "Number of Requests"

    c4 = ws.cell(row=1, column=4)
    c4.value = "Completion Time (s)"

    c5 = ws.cell(row=1, column=5)
    c5.value = "Average Time per Request (s)"

    c6 = ws.cell(row=1, column=6)
    c6.value = "Requests per Second"

    return wb

def write_to_excel(self, name, url, num_requests, request_time, average):
    """Write data to excel sheet."""
    ws = self.wb.active

    row = ws.max_row + 1
    print(f"Current row: {row}")

    ws.cell(row=row, column=1).value = name

    ws.cell(row=row, column=2).value = url

    ws.cell(row=row, column=3).value = num_requests

    ws.cell(row=row, column=4).value = request_time

    ws.cell(row=row, column=5).value = average

    # requests per second should be 1 / average (1 because there is only 1 thread)
    ws.cell(row=row, column=6).value = 1 / average

    self.wb.save(PATH)